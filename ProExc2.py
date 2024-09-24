import os
import pandas as pd
import re
import unicodedata
from openpyxl import load_workbook

def procesar_y_pegar_datos_mensuales(ruta_carpeta, ubicacion):
    # Construir el nombre del archivo CSV
    ubicacion_sanitizada = re.sub(r'[<>:"/\\|?*]', '', ubicacion)
    ubicacion_sanitizada = ubicacion_sanitizada.replace(' ', '_')
    ubicacion_sanitizada = unicodedata.normalize('NFD', ubicacion_sanitizada).encode('ascii', 'ignore').decode('utf-8')
    nombre_archivo_csv = f"nasa_data_monthly_{ubicacion_sanitizada}.csv"
    ruta_archivo_csv = os.path.join(ruta_carpeta, nombre_archivo_csv)
    
    # Verificar si el archivo CSV existe
    if not os.path.exists(ruta_archivo_csv):
        print(f"El archivo {ruta_archivo_csv} no existe.")
        return

    # Leer el archivo CSV como texto
    with open(ruta_archivo_csv, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Procesar datos desde la fila 22 hasta la 32 (índices 21 a 31)
    data_lines_T2M = lines[21:32]
    T2M_data = []
    for line in data_lines_T2M:
        line = line.strip()
        # Dividir por comas
        row = line.split(',')
        T2M_data.append(row)

    # Procesar datos desde la fila 33 hasta la 43 (índices 32 a 43)
    data_lines_SW_DWN = lines[32:43]
    SW_DWN_data = []
    for line in data_lines_SW_DWN:
        line = line.strip()
        # Dividir por comas
        row = line.split(',')
        SW_DWN_data.append(row)

    # Convertir las listas en DataFrames
    df_T2M = pd.DataFrame(T2M_data)
    df_SW_DWN = pd.DataFrame(SW_DWN_data)

    # Definir la ruta del archivo Excel de destino
    ruta_excel_destino = r'C:\Users\adamonte\Desktop\AGUSTIN\Python\NASA AUTOM Prueba - Modelo Gen ee.xlsx'

    # Cargar el libro de trabajo existente
    wb = load_workbook(ruta_excel_destino)
    sheet_name = 'NASALarc'
    if sheet_name not in wb.sheetnames:
        print(f"La hoja '{sheet_name}' no existe en el archivo Excel.")
        return
    ws = wb[sheet_name]

    # Escribir T2M a partir de la fila 26
    start_row_T2M = 26  # Fila 26
    start_col = 1       # Columna A (1-indexado)

    for idx, row in enumerate(df_T2M.values):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=start_row_T2M + idx, column=start_col + col_idx, value=value)
            # Intentar convertir el valor a número si es posible
            try:
                num_value = float(value)
                cell.value = num_value
                cell.number_format = '#,##0.00'
            except ValueError:
                pass  # Mantener el valor original si no es un número

    # Escribir ALLSKY_SFC_SW_DWN a partir de la fila 12
    start_row_SW_DWN = 12  # Fila 12

    for idx, row in enumerate(df_SW_DWN.values):
        for col_idx, value in enumerate(row):
            cell = ws.cell(row=start_row_SW_DWN + idx, column=start_col + col_idx, value=value)
            # Intentar convertir el valor a número si es posible
            try:
                num_value = float(value)
                cell.value = num_value
                cell.number_format = '#,##0.00'
            except ValueError:
                pass  # Mantener el valor original si no es un número

    # Guardar el libro de trabajo
    wb.save(ruta_excel_destino)

    print(f"Datos mensuales procesados y pegados en el archivo Excel en la hoja '{sheet_name}'.")

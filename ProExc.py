import os
import pandas as pd
import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers  # Importar para formatos numéricos

def procesar_y_pegar_datos_horarios(ruta_carpeta, ubicacion):
    # Construir el nombre del archivo CSV
    ubicacion_sanitizada = re.sub(r'[<>:"/\\|?*]', '', ubicacion)
    ubicacion_sanitizada = ubicacion_sanitizada.replace(' ', '_')
    ubicacion_sanitizada = unicodedata.normalize('NFD', ubicacion_sanitizada).encode('ascii', 'ignore').decode('utf-8')
    nombre_archivo_csv = f"nasa_data_hourly_{ubicacion_sanitizada}.csv"
    ruta_archivo_csv = os.path.join(ruta_carpeta, nombre_archivo_csv)
    
    # Verificar si el archivo CSV existe
    if not os.path.exists(ruta_archivo_csv):
        print(f"El archivo {ruta_archivo_csv} no existe.")
        return

    # Leer el archivo CSV como texto
    with open(ruta_archivo_csv, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Saltar las primeras 20 líneas (hasta la fila 20)
    data_lines = lines[20:]

    # Procesar cada línea, dividir por comas y convertir elementos a números si es posible
    data = []
    for line in data_lines:
        # Remover caracteres de nueva línea y espacios en blanco
        line = line.strip()
        # Dividir por comas
        row = line.split(',')
        processed_row = []
        for element in row:
            # Intentar convertir el elemento a número (float)
            try:
                value = float(element)
            except ValueError:
                # Si no es posible, mantener el valor original (por ejemplo, cadenas de texto)
                value = element
            processed_row.append(value)
        data.append(processed_row)

    # Convertir la lista de listas en un DataFrame de pandas
    df = pd.DataFrame(data)

    # Definir la ruta del archivo Excel de destino
    ruta_excel_destino = r'C:\Users\adamonte\Desktop\AGUSTIN\Python\NASA AUTOM Prueba - Modelo Gen ee.xlsx'

    # Escribir los datos en la hoja 'NASALarc' a partir de la celda R11
    sheet_name = 'NASALarc'
    start_row = 11  # R11 es la fila 11 (1-indexado)
    start_col = 18  # Columna R es la columna 18 (1-indexado)

    # Cargar el libro de trabajo existente
    wb = load_workbook(ruta_excel_destino)
    if sheet_name not in wb.sheetnames:
        print(f"La hoja {sheet_name} no existe en el archivo Excel.")
        return

    ws = wb[sheet_name]

    # Convertir el DataFrame en filas que openpyxl puede escribir
    rows = dataframe_to_rows(df, index=False, header=False)

    # Escribir los datos en la hoja
    for r_idx, row in enumerate(rows, start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Si el valor es numérico, aplicar formato numérico
            if isinstance(value, (int, float)):
                # Aplicar un formato numérico con dos decimales y separador decimal de coma
                cell.number_format = '#,##0.00'

    # Guardar el libro de trabajo
    wb.save(ruta_excel_destino)

    print(f"Datos procesados y pegados en el archivo Excel en la hoja '{sheet_name}' a partir de la celda R11.")